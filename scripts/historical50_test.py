#!/usr/bin/env python3
"""
ClaimVelo Historical 50-Flight Test — STANDALONE DATA TEST.

No production code changes. Uses the EXACT same provider fetch logic and
Rules Engine decision path as supabase/functions/_shared/evaluate.ts.

For each of the 50 flights in the Excel file:
  1. Calls AeroDataBox and AviationStack APIs (same endpoints, same parsing).
  2. Cross-checks flight number / date / origin / destination.
  3. Applies the Rules Engine decision logic (jurisdiction, statute of
     limitations, delay threshold, compensation calculation).
  4. Reports per-flight results + aggregate summary.

Outputs:
  - LIKELY_ELIGIBLE / LIKELY_NOT_ELIGIBLE / MANUAL_REVIEW / INSUFFICIENT_FLIGHT_DATA
  - HISTORICAL_DATA_UNAVAILABLE when a provider returns no data
  - PROVIDER_QUOTA_EXHAUSTED when quota is hit
"""

import json
import math
import os
import sys
import time
import urllib.request
import urllib.parse
import urllib.error
import xml.etree.ElementTree  # noqa
from datetime import datetime, timezone

import openpyxl

# ── Load API keys from the platform-managed env file ──────────────────────────
ENV_FILE = "/run/base44/app.env"
API_KEYS = {}
if os.path.exists(ENV_FILE):
    with open(ENV_FILE) as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                API_KEYS[k.strip()] = v.strip()

AERODATABOX_KEY = API_KEYS.get("AERODATABOX_API_KEY", "")
AVIATIONSTACK_KEY = API_KEYS.get("AVIATIONSTACK_API_KEY", "")

if not AERODATABOX_KEY:
    print("WARNING: AERODATABOX_API_KEY not found", file=sys.stderr)
if not AVIATIONSTACK_KEY:
    print("WARNING: AVIATIONSTACK_API_KEY not found", file=sys.stderr)

# ── Quota tracking ─────────────────────────────────────────────────────────────
aero_quota_exhausted = False
avia_quota_exhausted = False
aero_calls = 0
avia_calls = 0
aero_errors = 0
avia_errors = 0

# ── Helpers (replicated from evaluate.ts) ─────────────────────────────────────

def normalize_flight_number(s):
    return "".join(c for c in (s or "") if c.isalnum()).upper()

def normalize_date(s):
    if s is None:
        return ""
    if isinstance(s, datetime):
        return s.strftime("%Y-%m-%d")
    return str(s)[:10]

def normalize_iata(s):
    return (s or "").strip().upper()

# ── Airport code sets ──────────────────────────────────────────────────────────

UK_AIRPORT_CODES = {
    "LHR","LGW","STN","LTN","LCY","SEN","MAN","EDI","BHX","GLA","BRS","NCL","ABZ",
    "LPL","EMA","LBA","CWL","BFS","SOU","EXT","NWI","INV","JER","GCI","IOM",
}

EU_EEA_AIRPORT_CODES = {
    "CDG","ORY","NCE","LYS","MRS","TLS","BOD","BIA","NTE","MPL",
    "FRA","MUC","DUS","HAM","BER","CGN","STR","HAJ","LEJ","DRS","NUE",
    "AMS","RTM","EIN","BRU","CRL","LUX",
    "FCO","MXP","LIN","BGY","VCE","NAP","CIA","FLR","TRN","BLQ","CTA","PSA","BRI","CAG","PMO","TSF",
    "MAD","BCN","VLC","AGP","PMI","SVQ","BIO","OVD","SCQ","TFN","TFS","LPA","ACE","FUE","ALC","GRX","XRY",
    "LIS","OPO","FAO","FNC","PDL",
    "ATH","SKG","HER","RHO","CHQ","KGS","CFU","JKH","VOL","PVK","AOI",
    "VIE","SZG","INN","GRZ","LNZ","HOH",
    "ZRH","GVA","BSL","BRN","LUG",
    "CPH","RNN","BLL","AAL","AAR",
    "ARN","BMA","GOT","MMX","NBQ","LLA","UME","OSD","VBY","KLR","RNB","GEV",
    "OSL","BGO","TRD","SVG","TOS","KKN","BOO","HAU","AES","EVE","BNN","FRO",
    "HEL","TMP","TKU","OUL","KUO","JYV","SVL","KEM","MIK","RVN","KTT",
    "TLL","TAY",
    "RIX","VSI","LPX","VNO","KUN","PLQ","SQQ",
    "DUB","ORK","SNN","NOC","KIR","WAT",
    "KEF","REK","AEY","IFJ","GRM","HFN","HUS","THO","VEY",
    "WAW","KRK","GDN","KTW","WRO","POZ","LCJ","RZE","SZZ","BZG",
    "PRG","BRQ","OSR","KLV",
    "BUD","DEB",
    "OTP","CLJ","TSR","CNR","SBZ","IAS","ARW","BCM","CMB","CSB","CRA","ISL","SUJ",
    "SOF","VAR","BOJ","PDV","GSB",
    "ZAG","SPU","DBV","RJK","OSI","ZAD","PUY","BWK",
    "LJU","MBX","POW",
    "BTS","KSC","TAT","DSV",
    "MLA","GZM",
    "LCA","PFO",
    "IST","SAW","AYT","ADB","ESB",
}

ISRAELI_AIRPORT_CODES = {"TLV","BGW","ETM","HFA","SDV","VDA","KCN"}

BRAZIL_AIRPORT_CODES = {
    "GRU","GIG","BSB","CGH","SDU","CNF","POA","REC","SSA","FOR","CWB","MAO","BEL",
    "GYN","VCP","FLN","NAT","MCZ","VIX","CGB","SLZ","UDI","RAO","ATM","JPA","MCP",
    "PVH","STM","BPS","MGB","THE","MCZ",
}

UK_CARRIERS = {"BA","VS","U2","LS","BE","GR","WQ","T3","JD","EX","MM"}

EU_CARRIERS = {
    "LH","LX","OS","SN","EN","DE","EW","4U","AB",
    "AF","U2","A5","TO","SS","XL","BJ",
    "KL","HV","WA",
    "IB","UX","FR","VY","QS",
    "EI","WI","FR",
    "AZ","NO","IG","EI","VE","XR","W6",
    "SK","DY","W6","FI","RC","EF",
    "TP","S4",
    "A3","OA","EG","W6",
    "LO","W6","BT","RJ",
    "BT","LO","RJ","PS",
    "OK","QS",
    "MA","W6",
    "RO","W6","0B",
    "FB","W6",
    "OU",
    "JP",
    "W6","OK",
    "KM",
    "CY","W6",
    "TK","PC",
    "FI",
    "AY","W6",
    "OS","W6",
}

BRAZIL_CARRIERS = {"JJ","LA","G3","AD","RJ","2Z","O6","W3"}

AIRPORT_COORDS = {
    "TLV":[32.011,34.887],"SDV":[32.419,34.880],"ETM":[29.698,35.013],"VDA":[29.569,35.009],"KCN":[29.632,35.014],
    "LHR":[51.477,-0.461],"LGW":[51.148,-0.190],"STN":[51.885,0.235],"LTN":[51.874,-0.368],"LCY":[51.505,0.055],
    "MAN":[53.354,-2.275],"EDI":[55.950,-3.373],"BHX":[52.453,-1.748],"GLA":[55.872,-4.433],"BRS":[51.382,-2.719],
    "CDG":[49.009,2.548],"ORY":[48.724,2.380],"NCE":[43.658,7.215],"LYS":[45.726,5.081],"MRS":[43.435,5.215],
    "AMS":[52.308,4.764],"BRU":[50.902,4.484],
    "FRA":[50.033,8.570],"MUC":[48.354,11.786],"BER":[52.366,13.503],"DUS":[51.289,6.767],"HAM":[53.630,10.006],
    "MAD":[40.472,-3.561],"BCN":[41.297,2.078],"PMI":[39.551,2.739],"AGP":[36.675,-4.499],
    "FCO":[41.800,12.239],"MXP":[45.630,8.723],"LIN":[45.445,9.277],"VCE":[45.505,12.352],"NAP":[40.886,14.291],
    "LIS":[38.781,-9.136],"OPO":[41.248,-8.681],
    "ATH":[37.936,23.944],"SKG":[40.520,22.971],
    "VIE":[48.110,16.570],"ZRH":[47.458,8.548],"GVA":[46.238,6.109],
    "CPH":[55.618,12.656],"ARN":[59.651,17.919],"OSL":[60.194,11.100],"HEL":[60.317,24.963],
    "DUB":[53.421,-6.270],"SNN":[52.702,-8.925],"KEF":[63.985,-22.606],
    "WAW":[52.165,20.967],"PRG":[50.100,14.260],"BUD":[47.433,19.261],
    "OTP":[44.572,26.102],"SOF":[42.696,23.411],"ZAG":[45.743,16.069],
    "RIX":[56.924,23.971],"TLL":[59.413,24.832],"VNO":[54.634,25.285],
    "JFK":[40.640,-73.779],"LAX":[33.943,-118.408],"ORD":[41.978,-87.905],"ATL":[33.640,-84.427],
    "DXB":[25.253,55.363],
    "GRU":[-23.432,-46.469],"GIG":[-22.808,-43.244],"BSB":[-15.869,-47.920],"CNF":[-19.633,-43.968],
    "DOH":[25.273,51.608],"MUC2":[0,0],  # placeholder
}

# Add MUC properly
AIRPORT_COORDS["MUC"] = [48.354, 11.786]

EXTRAORDINARY_REASONS = {"WEATHER","ATC","SECURITY","STRIKE"}

REASON_MAP = {
    "carrier":"CARRIER","technical":"TECHNICAL","crew":"CREW",
    "weather":"WEATHER","atc":"ATC","air traffic control":"ATC",
    "security":"SECURITY","strike":"STRIKE",
}

MIN_DELAY_EU_UK = 180
MIN_DELAY_IL = 480
KM_SHORT = 1500
KM_MEDIUM = 3500

COMPENSATION = {
    "EU261": {"short":{"full":250,"reduced":125},"medium":{"full":400,"reduced":200},"long":{"full":600,"reduced":300}},
    "UK261": {"short":{"full":220,"reduced":110},"medium":{"full":350,"reduced":175},"long":{"full":520,"reduced":260}},
    "ISRAEL": {"short":{"full":1470,"reduced":1470},"medium":{"full":2390,"reduced":2390},"long":{"full":3530,"reduced":3530}},
}

REDUCTION_THRESHOLDS = {"short":120,"medium":180,"long":240}

def classify_reason(raw):
    lower = (raw or "").lower().strip()
    for key, code in REASON_MAP.items():
        if key in lower:
            return code
    return "CARRIER"

def is_israeli_route(dep, arr):
    return dep.upper() in ISRAELI_AIRPORT_CODES or arr.upper() in ISRAELI_AIRPORT_CODES

def is_uk_airport(code):
    return code.upper() in UK_AIRPORT_CODES

def is_eu_airport(code):
    return code.upper() in EU_EEA_AIRPORT_CODES

def is_brazilian_route(dep, arr):
    return dep.upper() in BRAZIL_AIRPORT_CODES or arr.upper() in BRAZIL_AIRPORT_CODES

def is_uk_carrier(iata):
    return iata and iata.upper() in UK_CARRIERS

def is_eu_carrier(iata):
    return iata and iata.upper() in EU_CARRIERS

def carrier_country(iata):
    if not iata: return "unknown"
    u = iata.upper()
    if u in UK_CARRIERS: return "UK"
    if u in EU_CARRIERS: return "EU"
    if u in BRAZIL_CARRIERS: return "BR"
    return "non_eu"

def determine_jurisdiction(dep, arr, operating_carrier):
    d = dep.upper()
    a = arr.upper()
    if is_brazilian_route(d, a):
        return ("BRAZIL_REVIEW", "Brazilian route — manual review required.")
    if is_israeli_route(d, a):
        return ("ISRAEL", "Israeli Aviation Services Law applies.")
    if is_uk_airport(d):
        return ("UK261", "UK261 applies (departure from UK airport).")
    if is_eu_airport(d):
        return ("EU261", "EU261 applies (departure from EU/EEA airport).")
    if is_uk_airport(a):
        if is_uk_carrier(operating_carrier):
            return ("UK261", "UK261 applies (UK-licensed carrier arriving at UK).")
        if carrier_country(operating_carrier) == "unknown":
            return ("NONE", "Cannot determine operating carrier country for UK arrival route.")
        return ("NONE", "UK261 does not apply (non-UK carrier on non-UK→UK route).")
    if is_eu_airport(a):
        if is_eu_carrier(operating_carrier):
            return ("EU261", "EU261 applies (EU-licensed carrier arriving at EU).")
        if carrier_country(operating_carrier) == "unknown":
            return ("NONE", "Cannot determine operating carrier country for EU arrival route.")
        return ("NONE", "EU261 does not apply (non-EU carrier on non-EU→EU route).")
    return ("NONE", "Route not covered by EU261/UK261/Israeli/Brazilian regulations.")

def years_between(from_date, to_date):
    y = to_date.year - from_date.year
    m = to_date.month - from_date.month
    if m < 0 or (m == 0 and to_date.day < from_date.day):
        y -= 1
    return y

def haversine_km(a, b):
    R = 6371
    d_lat = math.radians(b[0] - a[0])
    d_lon = math.radians(b[1] - a[1])
    lat1 = math.radians(a[0])
    lat2 = math.radians(b[0])
    h = math.sin(d_lat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(d_lon/2)**2
    return round(2 * R * math.asin(math.sqrt(h)))

def get_distance_category(dep, arr):
    ca = AIRPORT_COORDS.get(dep.upper())
    cb = AIRPORT_COORDS.get(arr.upper())
    if not ca or not cb: return None, None
    km = haversine_km(ca, cb)
    if km <= KM_SHORT: return "short", km
    if km <= KM_MEDIUM: return "medium", km
    return "long", km

def calc_delay_compensation(dep, arr, delay_min, jurisdiction):
    cat, dist_km = get_distance_category(dep, arr)
    if not cat: return None
    jur_key = jurisdiction
    if jur_key not in COMPENSATION:
        jur_key = "EU261"
    comp = COMPENSATION[jur_key][cat]
    currency = "GBP" if jurisdiction == "UK261" else "ILS" if jurisdiction == "ISRAEL" else "EUR"
    if cat == "short":
        amount = comp["full"]
    elif cat == "medium":
        amount = comp["full"]
    else:
        if delay_min >= MIN_DELAY_EU_UK and delay_min < REDUCTION_THRESHOLDS["long"]:
            amount = comp["reduced"]
        else:
            amount = comp["full"]
    if jurisdiction == "ISRAEL":
        amount = comp["full"]
    return {"amount": amount, "currency": currency, "distance_km": dist_km}

def currency_symbol(currency):
    if currency == "ILS": return "₪"
    if currency == "GBP": return "£"
    return "€"

# ── Provider fetch (exact replica of evaluate.ts) ──────────────────────────────

def fetch_aerodatabox(fn, date_str):
    """Replicates fetchAeroDataBox from evaluate.ts exactly."""
    global aero_quota_exhausted, aero_calls, aero_errors
    if aero_quota_exhausted:
        return None, "PROVIDER_QUOTA_EXHAUSTED"
    if not AERODATABOX_KEY or not fn or not date_str:
        return None, "NO_KEY"
    iata = normalize_flight_number(fn)
    url = f"https://aerodatabox.p.rapidapi.com/flights/number/{iata}/{date_str}"
    aero_calls += 1
    try:
        req = urllib.request.Request(url, headers={
            "x-rapidapi-host": "aerodatabox.p.rapidapi.com",
            "x-rapidapi-key": AERODATABOX_KEY,
        })
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status != 200:
                return None, f"HTTP_{resp.status}"
            data = json.loads(resp.read().decode())
            if not isinstance(data, list) or not data:
                return None, "EMPTY"
            flights = []
            for f in data:
                dep = f.get("departure", {}) or {}
                arr = f.get("arrival", {}) or {}
                airline = f.get("airline", {}) or {}
                dep_airport = dep.get("airport", {}) or {}
                arr_airport = arr.get("airport", {}) or {}
                dep_sched = dep.get("scheduledTime", {}) or {}
                arr_sched = arr.get("scheduledTime", {}) or {}
                sched_dep = dep_sched.get("utc") or dep_sched.get("local")
                sched_arr = arr_sched.get("utc") or arr_sched.get("local")
                dep_runway = dep.get("runwayTime", {}).get("utc") if dep.get("runwayTime") else None
                dep_runway_local = dep.get("runwayTime", {}).get("local") if dep.get("runwayTime") else None
                dep_runway = dep_runway or dep_runway_local
                dep_revised = dep.get("revisedTime", {}).get("utc") if dep.get("revisedTime") else None
                dep_revised_local = dep.get("revisedTime", {}).get("local") if dep.get("revisedTime") else None
                dep_revised = dep_revised or dep_revised_local
                arr_runway = arr.get("runwayTime", {}).get("utc") if arr.get("runwayTime") else None
                arr_runway_local = arr.get("runwayTime", {}).get("local") if arr.get("runwayTime") else None
                arr_runway = arr_runway or arr_runway_local
                arr_revised = arr.get("revisedTime", {}).get("utc") if arr.get("revisedTime") else None
                arr_revised_local = arr.get("revisedTime", {}).get("local") if arr.get("revisedTime") else None
                arr_revised = arr_revised or arr_revised_local
                actual_dep = dep_runway or (dep_revised if dep_revised and dep_revised != sched_dep else None)
                actual_arr = arr_runway or (arr_revised if arr_revised and arr_revised != sched_arr else None)
                delay_min = None
                if sched_arr and actual_arr:
                    delay_min = max(0, round((datetime.fromisoformat(actual_arr.replace("Z","+00:00")).timestamp() -
                                             datetime.fromisoformat(sched_arr.replace("Z","+00:00")).timestamp()) / 60))
                flights.append({
                    "flightNumber": normalize_flight_number(f.get("number") or fn),
                    "flightDate": normalize_date(date_str),
                    "origin": normalize_iata(dep_airport.get("iata") or ""),
                    "destination": normalize_iata(arr_airport.get("iata") or ""),
                    "scheduledDeparture": sched_dep,
                    "scheduledArrival": sched_arr,
                    "actualDeparture": actual_dep,
                    "actualArrival": actual_arr,
                    "delayMinutes": delay_min,
                    "status": f.get("status") or "scheduled",
                    "operatingCarrier": airline.get("iata"),
                    "operatingCarrierName": airline.get("name"),
                    "marketingCarrier": None,
                    "codeshareStatus": f.get("codeshareStatus"),
                })
            return {"source": "aerodatabox", "flights": flights, "raw_count": len(data)}, "OK"
    except urllib.error.HTTPError as e:
        aero_errors += 1
        if e.code == 429:
            aero_quota_exhausted = True
            return None, "PROVIDER_QUOTA_EXHAUSTED"
        return None, f"HTTP_{e.code}"
    except Exception as e:
        aero_errors += 1
        return None, f"ERROR: {str(e)[:100]}"

def fetch_aviationstack(fn, date_str):
    """Replicates fetchAviationStack from evaluate.ts exactly."""
    global avia_quota_exhausted, avia_calls, avia_errors
    if avia_quota_exhausted:
        return None, "PROVIDER_QUOTA_EXHAUSTED"
    if not AVIATIONSTACK_KEY or not fn or not date_str:
        return None, "NO_KEY"
    iata = normalize_flight_number(fn)
    params = urllib.parse.urlencode({"access_key": AVIATIONSTACK_KEY, "flight_iata": iata, "flight_date": date_str})
    url = f"http://api.aviationstack.com/v1/flights?{params}"
    avia_calls += 1
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = json.loads(resp.read().decode())
            if raw.get("error"):
                # Check for quota error
                err_ctx = raw.get("error", {})
                if isinstance(err_ctx, dict):
                    err_code = err_ctx.get("code", "")
                    if "limit" in str(err_code).lower() or "quota" in str(err_code).lower():
                        avia_quota_exhausted = True
                        return None, "PROVIDER_QUOTA_EXHAUSTED"
                return None, f"API_ERROR: {str(err_ctx)[:100]}"
            data = raw.get("data")
            if not data or not data:
                return None, "EMPTY"
            flights = []
            for f in data:
                dep = f.get("departure", {}) or {}
                arr = f.get("arrival", {}) or {}
                flight_info = f.get("flight", {}) or {}
                airline = f.get("airline", {}) or {}
                sched_dep = dep.get("scheduled")
                actual_dep = dep.get("actual") or dep.get("estimated")
                sched_arr = arr.get("scheduled")
                actual_arr = arr.get("actual") or arr.get("estimated")
                delay_min = None
                if sched_arr and actual_arr:
                    try:
                        delay_min = max(0, round((datetime.fromisoformat(actual_arr.replace("Z","+00:00")).timestamp() -
                                                  datetime.fromisoformat(sched_arr.replace("Z","+00:00")).timestamp()) / 60))
                    except:
                        pass
                elif dep.get("delay"):
                    delay_min = int(dep["delay"]) if dep["delay"] else None
                codeshared = flight_info.get("codeshared") or {}
                marketing_flight_iata = codeshared.get("flight_iata") if codeshared else None
                flights.append({
                    "flightNumber": normalize_flight_number(marketing_flight_iata or flight_info.get("iata") or fn),
                    "flightDate": normalize_date(f.get("flight_date") or date_str),
                    "origin": normalize_iata(dep.get("iata") or ""),
                    "destination": normalize_iata(arr.get("iata") or ""),
                    "scheduledDeparture": sched_dep,
                    "scheduledArrival": sched_arr,
                    "actualDeparture": actual_dep,
                    "actualArrival": actual_arr,
                    "delayMinutes": delay_min,
                    "status": f.get("flight_status") or "scheduled",
                    "operatingCarrier": airline.get("iata"),
                    "operatingCarrierName": airline.get("name"),
                    "marketingCarrier": codeshared.get("airline_iata") if codeshared else None,
                    "codeshareStatus": "IsCodeshared" if codeshared else "IsOperator",
                })
            return {"source": "aviationstack", "flights": flights, "raw_count": len(data)}, "OK"
    except urllib.error.HTTPError as e:
        avia_errors += 1
        if e.code == 429:
            avia_quota_exhausted = True
            return None, "PROVIDER_QUOTA_EXHAUSTED"
        return None, f"HTTP_{e.code}"
    except Exception as e:
        avia_errors += 1
        return None, f"ERROR: {str(e)[:100]}"

# ── Cross-check (replica of crossCheck from evaluate.ts) ───────────────────────

def cross_check(claim_fn, claim_date, claim_origin, claim_dest, providers):
    c_fn = normalize_flight_number(claim_fn)
    c_date = normalize_date(claim_date)
    c_origin = normalize_iata(claim_origin)
    c_dest = normalize_iata(claim_dest)

    if not providers or all(len(p["flights"]) == 0 for p in providers):
        return {"status": "no_data", "matched": None, "details": {"reason": "No flight data returned by any provider"}}

    all_flights = []
    for p in providers:
        all_flights.extend(p["flights"])

    all_matches = [f for f in all_flights if
                   normalize_flight_number(f["flightNumber"]) == c_fn and
                   normalize_date(f["flightDate"]) == c_date and
                   normalize_iata(f["origin"]) == c_origin and
                   normalize_iata(f["destination"]) == c_dest]

    is_cancelled = lambda s: s and s.lower() in ("cancelled", "canceled")
    matched = None
    for f in all_matches:
        if is_cancelled(f["status"]):
            matched = f
            break
    if not matched:
        for f in all_matches:
            if f["actualArrival"]:
                matched = f
                break
    if not matched and all_matches:
        matched = all_matches[0]

    if not matched:
        ref = all_flights[0] if all_flights else None
        mismatches = []
        if ref:
            if normalize_flight_number(ref["flightNumber"]) != c_fn: mismatches.append("flight number")
            if normalize_date(ref["flightDate"]) != c_date: mismatches.append("flight date")
            if normalize_iata(ref["origin"]) != c_origin: mismatches.append("origin")
            if normalize_iata(ref["destination"]) != c_dest: mismatches.append("destination")
        return {"status": "mismatch", "matched": None, "details": {"reason": f"Mismatch on: {', '.join(mismatches) or 'no candidates'}"}}

    # Provider conflict on delay
    provider_conflict = False
    matched_per_provider = []
    for p in providers:
        for f in p["flights"]:
            if (normalize_flight_number(f["flightNumber"]) == c_fn and
                normalize_date(f["flightDate"]) == c_date and
                normalize_iata(f["origin"]) == c_origin and
                normalize_iata(f["destination"]) == c_dest):
                matched_per_provider.append(f)
    if len(matched_per_provider) >= 2:
        with_delay = [f for f in matched_per_provider if f["delayMinutes"] is not None and f["actualArrival"]]
        if len(with_delay) >= 2:
            delays = [f["delayMinutes"] for f in with_delay]
            if max(delays) - min(delays) > 10:
                provider_conflict = True

    # Carrier conflict
    carrier_conflict = False
    if len(matched_per_provider) >= 2:
        carriers = [f["operatingCarrier"] for f in matched_per_provider if f["operatingCarrier"]]
        unique = set(c.upper() for c in carriers)
        if len(unique) > 1:
            carrier_conflict = True

    status = "carrier_conflict" if carrier_conflict else ("conflict" if provider_conflict else "matched")
    return {"status": status, "matched": matched, "details": {"provider_conflict": provider_conflict, "carrier_conflict": carrier_conflict}}

# ── Rules Engine decision (replica of evaluateClaimInternal decision path) ────

def run_rules_engine(flight, cc_result, providers):
    """
    Runs the Rules Engine decision path from evaluate.ts.
    Returns: (decision, delay_minutes, reason, compensation, jurisdiction, source)
    """
    dep = normalize_iata(flight["Origin_IATA"])
    arr = normalize_iata(flight["Destination_IATA"])
    fn = flight["Marketing_Flight_Number"]
    date_str = normalize_date(flight["Departure_DateTime"])
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    primary_source = "none"
    if cc_result["matched"]:
        aero_prov = next((p for p in providers if p["source"] == "aerodatabox"), None)
        if aero_prov and cc_result["matched"] in aero_prov["flights"]:
            primary_source = "aerodatabox"
        else:
            primary_source = "aviationstack"
    elif providers:
        primary_source = providers[0]["source"]

    # Stage 1: Statute of limitations
    if date_str:
        dep_date = datetime.fromisoformat(date_str)
        age_years = years_between(dep_date, now)
        if is_israeli_route(dep, arr) and age_years > 4:
            return ("LIKELY_NOT_ELIGIBLE", None, f"Israeli route older than 4 years ({age_years}y) — statute of limitations exceeded.", None, "NONE", None)
        if (is_uk_airport(dep) or is_eu_airport(dep) or is_uk_airport(arr) or is_eu_airport(arr)) and age_years > 6:
            return ("LIKELY_NOT_ELIGIBLE", None, f"EU/UK route older than 6 years ({age_years}y) — statute of limitations exceeded.", None, "NONE", None)

    # Stage 1b: Brazil
    if is_brazilian_route(dep, arr):
        return ("MANUAL_REVIEW", None, "Brazilian route — manual review required (ANAC rules not yet automated).", None, "BRAZIL_REVIEW", None)

    # Stage 1c: Non-covered route
    if not is_uk_airport(dep) and not is_eu_airport(dep) and not is_israeli_route(dep, arr) and not is_uk_airport(arr) and not is_eu_airport(arr):
        return ("LIKELY_NOT_ELIGIBLE", None, "Route not covered by EU261/UK261/Israeli/Brazilian regulations.", None, "NONE", None)

    # Stage 3: Cross-check
    if cc_result["status"] == "no_data":
        return ("INSUFFICIENT_FLIGHT_DATA", None, "No flight data available from any provider.", None, "NONE", primary_source if primary_source != "none" else None)
    if cc_result["status"] == "mismatch":
        return ("MANUAL_REVIEW", None, f"Cross-check failed: {cc_result['details']['reason']}", None, "NONE", primary_source)
    if cc_result["status"] == "conflict":
        return ("MANUAL_REVIEW", None, "Providers returned conflicting delay data.", None, "NONE", primary_source)
    if cc_result["status"] == "carrier_conflict":
        return ("MANUAL_REVIEW", None, "Providers disagree on the operating carrier — cannot determine jurisdiction.", None, "NONE", primary_source)

    matched = cc_result["matched"]

    # Stage 3d: Jurisdiction
    operating_carrier = matched["operatingCarrier"]
    jur, jur_detail = determine_jurisdiction(dep, arr, operating_carrier)

    if jur == "NONE":
        if "Cannot determine" in jur_detail:
            return ("MANUAL_REVIEW", None, jur_detail, None, "NONE", primary_source)
        return ("LIKELY_NOT_ELIGIBLE", matched["delayMinutes"], jur_detail, None, "NONE", primary_source)

    # Stage 4: Cancellation detection
    if matched["status"] and matched["status"].lower() in ("cancelled", "canceled"):
        # No cancellation_notice_date in the test data → Pending Check
        return ("MANUAL_REVIEW", None, "Flight cancelled — cancellation notice date missing, cannot apply Article 5 rules.", None, jur, primary_source)

    # Stage 4b: Completeness — actual times required
    if not matched["actualArrival"] or not matched["scheduledArrival"] or matched["delayMinutes"] is None:
        return ("INSUFFICIENT_FLIGHT_DATA", matched["delayMinutes"], "Provider data incomplete: actual/scheduled arrival times unavailable.", None, jur, primary_source)

    delay_minutes = matched["delayMinutes"]
    reason_code = classify_reason("")  # No airline_reason in test data → defaults to CARRIER

    # Stage 5: Extraordinary circumstances — no airline_reason in test data, skip

    # Stage 6: Delay threshold
    threshold = MIN_DELAY_IL if jur == "ISRAEL" else MIN_DELAY_EU_UK
    if delay_minutes < threshold:
        return ("LIKELY_NOT_ELIGIBLE", delay_minutes, f"Delay of {delay_minutes}min is below the {'8-hour' if jur == 'ISRAEL' else '3-hour'} ({threshold}min) delay compensation threshold. (source: {primary_source})", None, jur, primary_source)

    # Stage 7: Eligible
    comp = calc_delay_compensation(dep, arr, delay_minutes, jur)
    if not comp:
        return ("MANUAL_REVIEW", delay_minutes, f"Airport coordinates unavailable for {dep}/{arr}.", None, jur, primary_source)

    sym = currency_symbol(comp["currency"])
    return ("LIKELY_ELIGIBLE", delay_minutes,
            f"Delay of {delay_minutes}min — carrier responsibility. Jurisdiction: {jur}. (source: {primary_source})",
            f"{sym}{comp['amount']} {comp['currency']}", jur, primary_source)

# ── Main test runner ───────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook("/tmp/historical50.xlsx", data_only=True)
    ws = wb["Historical_50"]

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = row[i] if i < len(row) else None
        rows.append(row_dict)

    print(f"Loaded {len(rows)} flights from Historical_50 sheet\n")
    print(f"{'#':>3}  {'Flight':<14} {'Date':<12} {'Route':<10} {'AeroDB':>7} {'AvStack':>7} {'CrossChk':<12} {'Decision':<25} {'Delay':>6} {'Comp':<12} {'Notes'}")
    print("=" * 160)

    results = []
    for idx, flight in enumerate(rows):
        fn = flight["Marketing_Flight_Number"]
        date_str = normalize_date(str(flight["Departure_DateTime"]))
        origin = normalize_iata(flight["Origin_IATA"])
        dest = normalize_iata(flight["Destination_IATA"])
        row_id = flight.get("Row_ID", "?")
        category = flight.get("Sample_Category", "")

        # Fetch providers
        aero_result, aero_status = fetch_aerodatabox(fn, date_str)
        time.sleep(0.3)  # rate limit courtesy
        avia_result, avia_status = fetch_aviationstack(fn, date_str)
        time.sleep(0.3)

        providers = []
        if aero_result:
            providers.append(aero_result)
        if avia_result:
            providers.append(avia_result)

        # Cross-check
        cc = cross_check(fn, date_str, origin, dest, providers)

        # Run Rules Engine
        decision, delay, reason, comp, jur, source = run_rules_engine(flight, cc, providers)

        # Determine verification level
        aero_found = aero_result is not None
        avia_found = avia_result is not None
        any_data = aero_found or avia_found
        matched_flight = cc["matched"]

        # Verification level
        if matched_flight and matched_flight["actualArrival"] and matched_flight["scheduledArrival"]:
            verification = "FULL"
        elif matched_flight and (matched_flight["actualArrival"] or matched_flight["scheduledArrival"]):
            verification = "PARTIAL"
        elif any_data:
            verification = "PARTIAL"
        else:
            verification = "NONE"

        # Provider status short codes
        aero_short = "DATA" if aero_found else ("QUOTA" if aero_status == "PROVIDER_QUOTA_EXHAUSTED" else "EMPTY")
        avia_short = "DATA" if avia_found else ("QUOTA" if avia_status == "PROVIDER_QUOTA_EXHAUSTED" else "EMPTY")

        cc_short = cc["status"][:12]
        delay_str = f"{delay}m" if delay is not None else "-"
        comp_str = comp or "-"

        # Detailed info
        detail_parts = []
        if matched_flight:
            if matched_flight["operatingCarrier"]:
                detail_parts.append(f"op={matched_flight['operatingCarrier']}")
            if matched_flight["status"]:
                detail_parts.append(f"st={matched_flight['status']}")
            if matched_flight["scheduledArrival"]:
                detail_parts.append(f"sched_arr={matched_flight['scheduledArrival'][:19]}")
            if matched_flight["actualArrival"]:
                detail_parts.append(f"act_arr={matched_flight['actualArrival'][:19]}")
        if not any_data:
            if aero_status == "PROVIDER_QUOTA_EXHAUSTED" or avia_status == "PROVIDER_QUOTA_EXHAUSTED":
                detail_parts.append("PROVIDER_QUOTA_EXHAUSTED")
            else:
                detail_parts.append("HISTORICAL_DATA_UNAVAILABLE")

        notes_str = " | ".join(detail_parts) if detail_parts else reason[:60]

        print(f"{idx+1:>3}  {fn:<14} {date_str:<12} {origin+'->'+dest:<10} {aero_short:>7} {avia_short:>7} {cc_short:<12} {decision:<25} {delay_str:>6} {comp_str:<12} {notes_str}")

        results.append({
            "idx": idx + 1,
            "row_id": row_id,
            "flight": fn,
            "date": date_str,
            "route": f"{origin}->{dest}",
            "category": category,
            "aero_found": aero_found,
            "avia_found": avia_found,
            "aero_status": aero_status,
            "avia_status": avia_status,
            "cross_check": cc["status"],
            "verification": verification,
            "matched": matched_flight is not None,
            "decision": decision,
            "delay_minutes": delay,
            "compensation": comp,
            "jurisdiction": jur,
            "source": source,
            "reason": reason,
            "operating_carrier": matched_flight["operatingCarrier"] if matched_flight else None,
            "flight_status": matched_flight["status"] if matched_flight else None,
            "scheduled_arrival": matched_flight["scheduledArrival"] if matched_flight else None,
            "actual_arrival": matched_flight["actualArrival"] if matched_flight else None,
        })

    # ── Aggregate summary ─────────────────────────────────────────────────────
    print("\n" + "=" * 160)
    print("AGGREGATE SUMMARY")
    print("=" * 160)

    total = len(results)
    fully_verified = sum(1 for r in results if r["verification"] == "FULL")
    partially_verified = sum(1 for r in results if r["verification"] == "PARTIAL")
    not_verified = sum(1 for r in results if r["verification"] == "NONE")

    likely_eligible = sum(1 for r in results if r["decision"] == "LIKELY_ELIGIBLE")
    likely_not_eligible = sum(1 for r in results if r["decision"] == "LIKELY_NOT_ELIGIBLE")
    manual_review = sum(1 for r in results if r["decision"] == "MANUAL_REVIEW")
    insufficient = sum(1 for r in results if r["decision"] == "INSUFFICIENT_FLIGHT_DATA")

    # Compensation total (where sufficiently supported)
    comp_total = {}
    for r in results:
        if r["decision"] == "LIKELY_ELIGIBLE" and r["compensation"]:
            # Parse "€250 EUR" etc
            parts = r["compensation"].split()
            if len(parts) >= 2:
                amt = int(parts[0].replace("₪","").replace("£","").replace("€",""))
                cur = parts[1]
                comp_total[cur] = comp_total.get(cur, 0) + amt

    print(f"\nTotal flights tested:           {total}")
    print(f"Fully verified (sched+actual):  {fully_verified}/{total}")
    print(f"Partially verified:             {partially_verified}/{total}")
    print(f"Could not be verified:          {not_verified}/{total}")
    print()
    print(f"LIKELY_ELIGIBLE:                {likely_eligible}")
    print(f"LIKELY_NOT_ELIGIBLE:            {likely_not_eligible}")
    print(f"MANUAL_REVIEW:                  {manual_review}")
    print(f"INSUFFICIENT_FLIGHT_DATA:       {insufficient}")
    print()
    print(f"Potential compensation total (where sufficiently supported):")
    for cur, amt in sorted(comp_total.items()):
        sym = "₪" if cur == "ILS" else "£" if cur == "GBP" else "€"
        print(f"  {sym}{amt:,} {cur}")
    if not comp_total:
        print("  (none — no flights reached LIKELY_ELIGIBLE with compensation)")
    print()
    print(f"Provider call stats:")
    print(f"  AeroDataBox: {aero_calls} calls, {aero_errors} errors, quota_exhausted={aero_quota_exhausted}")
    print(f"  AviationStack: {avia_calls} calls, {avia_errors} errors, quota_exhausted={avia_quota_exhausted}")
    print()

    # Per-category breakdown
    print("Per-category breakdown:")
    categories = {}
    for r in results:
        cat = r["category"] or "UNKNOWN"
        if cat not in categories:
            categories[cat] = {"total": 0, "verified": 0, "eligible": 0, "not_eligible": 0, "manual": 0, "insufficient": 0}
        categories[cat]["total"] += 1
        if r["verification"] in ("FULL", "PARTIAL"):
            categories[cat]["verified"] += 1
        if r["decision"] == "LIKELY_ELIGIBLE":
            categories[cat]["eligible"] += 1
        elif r["decision"] == "LIKELY_NOT_ELIGIBLE":
            categories[cat]["not_eligible"] += 1
        elif r["decision"] == "MANUAL_REVIEW":
            categories[cat]["manual"] += 1
        elif r["decision"] == "INSUFFICIENT_FLIGHT_DATA":
            categories[cat]["insufficient"] += 1

    for cat, stats in sorted(categories.items()):
        print(f"  {cat:<30} total={stats['total']:>2}  verified={stats['verified']:>2}  eligible={stats['eligible']:>2}  not_eligible={stats['not_eligible']:>2}  manual={stats['manual']:>2}  insufficient={stats['insufficient']:>2}")

    # Save detailed JSON
    with open("/tmp/historical50_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nDetailed JSON saved to /tmp/historical50_results.json")

if __name__ == "__main__":
    main()
